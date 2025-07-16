import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
import re
import os
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# def log_interaction(step_no, action, field, value=""):
#     file_path = os.path.join("reports", "reports.xlsx")

#     # Ensure the reports directory exists
#     os.makedirs(os.path.dirname(file_path), exist_ok=True)

#     # Load or create the Excel file
#     if os.path.exists(file_path):
#         wb = load_workbook(file_path)
#         ws = wb.active
#     else:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Test Log"
#         ws.append(["Step", "Action", "Field", "Value"])  # Header row

#     # Append the data
#     ws.append([step_no, action, field, value])

#     # Save the file
#     wb.save(file_path)


# def log_interaction(step_no, action, field, value="", status="Pass"):
#     file_path = os.path.join("reports", "reports.xlsx")
#     os.makedirs(os.path.dirname(file_path), exist_ok=True)

#     headers = ["Step", "Action", "Field", "Value", "Status"]

#     # Create or load workbook
#     if os.path.exists(file_path):
#         wb = load_workbook(file_path)
#         ws = wb.active
#         # Replace headers if needed
#         if [cell.value for cell in ws[1]] != headers:
#             ws.delete_rows(1)
#             ws.append(headers)
#     else:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Test Log"
#         ws.append(headers)
#     # if (status.lower() == "pass" and
#     #     (not str(step_no).strip()) and
#     #     (not str(action).strip()) and
#     #     (not str(field).strip()) and
#     #     (not str(value).strip())):
#     #     return
#     # Append row
#     ws.append([step_no, action, field, value, status])
#     status_cell = ws.cell(row=ws.max_row, column=5)

#     # ✅ Fill background and font color
#     if status.lower() == "pass":
#         status_cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")  # Green fill (ARGB)
#         status_cell.font = Font(color="FF006100", bold=True)  # Font color (ARGB)
#         status_cell.value = "Pass"
#     elif status.lower() == "fail":
#         status_cell.fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")  # Red fill (ARGB)
#         status_cell.font = Font(color="FF9C0006", bold=True)  # Font color (ARGB)
#         status_cell.value = "Fail"

#     wb.save(file_path)


def log_interaction(step_no, action, field, value="", status=""):

    file_path = os.path.join("reports", "reports.xlsx")
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    headers = ["Step", "Action", "Field", "Value", "Status"]
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        if [cell.value for cell in ws[1]] != headers:
            ws.delete_rows(1)
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Log"
        ws.append(headers)
    

    # ✅ Insert a blank row before a new test case starts
    if str(step_no).strip() == "1" and ws.max_row > 1:
        ws.append(["", "", "", "", ""])  # Blank row to separate test cases

    if (status.lower() == "pass" and
        (not str(step_no).strip()) and
        (not str(action).strip()) and
        (not str(field).strip()) and
        (not str(value).strip())):
        return  # Skip logging this row
    ws.append([step_no, action, field, value, status])
    status_cell = ws.cell(row=ws.max_row, column=5)

    if status.lower() == "invalid":
        status_cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        status_cell.font = Font(color="FF006100", bold=True)
        status_cell.value = "Validation failed , as expected value mismatched"
        print(f" ❌ Validation Failed ")
    elif "validate" in str(action).lower():
        status_cell.fill = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")  # light blue
        status_cell.font = Font(color="FF003366", bold=True)
        status_cell.value = "Validated"
       
        if value=="true":
            print(f"✅  Validation Successful : field is enabled")
        elif value=="false":
            print(f"✅  Validation Successful : field is read-only")
        else:
            print(f"✅  Validation Successful - {value}")

    elif status.lower() == "pass":
        status_cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        status_cell.font = Font(color="FF006100", bold=True)
        status_cell.value = "Pass"
    elif status.lower() == "fail":
        status_cell.fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        status_cell.font = Font(color="FF9C0006", bold=True)
        status_cell.value = "Fail"
    wb.save(file_path)
 



def js_click(driver, by, value, timeout=10):
    element = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((by, value)))
    driver.execute_script("arguments[0].click();", element)

def extract_navigation_steps(description):
    if not description or "Go to " not in description:
        return []

    # Extract the part after "Go to"
    path = description.split("Go to ", 1)[-1].strip()

    # Remove trailing period
    if path.endswith("."):
        path = path[:-1]

    # Split by '>' and strip extra spaces
    steps = [step.strip() for step in path.split(">")]

    return steps

def wait_and_click(driver, by, base_xpath, step_num = "",description="",timeout=10, enable_fallback=True):
    global last_failed_xpath
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, base_xpath))
        )
        aria_expanded = element.get_attribute("aria-expanded")

        if aria_expanded is None or aria_expanded.lower() == "false":
            WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, base_xpath)))
            ActionChains(driver).move_to_element(element).perform()
            element.click()
            if(description!="" and step_num != ""):
                log_interaction(step_num, "Click", description,"", "Pass")
            print(f"✅ Clicked element with base_xpath: {base_xpath}")
            return True
    except Exception as e:
        print(f"⚠️ Primary click failed on base_xpath: {base_xpath} - {str(e)}")
        last_failed_xpath = base_xpath
        take_screenshot_on_failure(driver)

        if enable_fallback and base_xpath:
            for i in range(1, 4):
                indexed_xpath = f"({base_xpath})[{i}]"
                try:
                    fallback_element = WebDriverWait(driver, timeout).until(
                        EC.element_to_be_clickable((by, indexed_xpath))
                    )
                    driver.execute_script("arguments[0].click();", fallback_element)
                    if(description!="" and step_num != ""):
                        log_interaction(step_num, "Click (fallback)", description)
                    return True
                except Exception as ex:
                    print(f"❌ Attempt {i} failed for xpath: {indexed_xpath} - {str(ex)}")
        
        print(f"❌ All attempts failed for base_xpath: {base_xpath}")
        take_screenshot_on_failure(driver)
        last_failed_xpath = base_xpath
        return False
        
def hover_on_an_element(driver, by, value, timeout=10):
    element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
    ActionChains(driver).move_to_element(element).perform()
    time.sleep(2)

def mouse_click(driver, by, value, timeout=10):
    element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
    actions = ActionChains(driver)
    actions.move_to_element(element).click().perform()

# Scroll into view just in case
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    time.sleep(0.3)

# Click using JavaScript
    driver.execute_script("arguments[0].click();", element)

# def wait_and_send_keys(driver, by, value, keys,timeout=20):
#     element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
#     ActionChains(driver).move_to_element(element).perform()
#     time.sleep(0.5)
#     element.click()
#     element.send_keys(keys)
#     time.sleep(1)
    # element.send_keys(Keys.RETURN)

def wait_and_send_keys(driver, by, value, keys, step_num = "" ,description="",timeout=20):
    global last_failed_xpath
    try:
        element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
        ActionChains(driver).move_to_element(element).perform()
        time.sleep(0.5)
        element.click()
        element.send_keys(keys)
        if(description != "" and step_num != ""):
            log_interaction(step_num, "Send Keys", description, keys, "Pass")
        time.sleep(1)
    except Exception as e:
        last_failed_xpath = value
        print(f"❌ Error in wait_and_send_keys at XPath: {value} - {str(e)}")
        take_screenshot_on_failure(driver)

def check_element_exist(driver, by, value, timeout=10):
    try:
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((by, value))
        )
        return True
    except TimeoutException:
        return False
    except Exception as e:
        print(f"Unexpected error in check_element_exist: {e}")
        return False

def checkInputExpanded(driver, by, value, timeout=10):
    try:
        element = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
        )
        aria_expanded = element.get_attribute("aria-expanded")
        if aria_expanded is not None:
            if aria_expanded.lower() == "false":
                return False
            else:
                return True
    except:
        return False

# def clear_input_field_and_send_keys(driver, by, value, keys, timeout=20):
#     try:
#         # Wait for the element to be clickable and store it
#         element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
#         time.sleep(0.5)  # Short buffer time before interaction
#         element.click()
#         element.send_keys(Keys.CONTROL + "a")
#         element.send_keys(Keys.DELETE)
#         element.send_keys(keys)
#         time.sleep(1)  # Give the page time to register input

#     except Exception as e:
#         print(f"Error in clear_input_field_and_send_keys: {e}")


def clear_input_field_and_send_keys(driver, by, value, keys, step_num="",description="",timeout=20):
    global last_failed_xpath
    try:
        element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
        time.sleep(0.5)  # Short buffer time before interaction
        element.click()
        element.send_keys(Keys.CONTROL + "a")
        element.send_keys(Keys.DELETE)
        element.send_keys(keys)
        if(description != "" and step_num != ""):
            log_interaction(step_num, "Send Keys", description, keys, "Pass")
        # base_test.steps_count += 1
        # print(f"✅ Successfully entered text at XPath: {value}")
        time.sleep(1)  # Give the page time to register input

    except Exception as e:
        last_failed_xpath = value
        print(f"❌ Error in clear_input_field_and_send_keys at XPath: {value} - {str(e)}")
        take_screenshot_on_failure(driver)


def no_of_elements_present(driver, by, value, timeout=20):
    try:
        elements = WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((by, value))
        )
        return len(elements)
    except TimeoutException:
        return 0

def get_locator(driver, by, value,):
    try:
        if(no_of_elements_present(driver, by, value)>1):
            print("Multiple elements found")
            return "("+value+")[1]"
        return value
    except TimeoutException:
        return None

def press_enter(driver, by, value, timeout=20):
    element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
    element.send_keys(Keys.ENTER)
    


def check_element_has_class(driver, by, locator, class_name, timeout=20):
    """
    Checks if the element's 'class' attribute contains a specific class name.

    Args:
        driver: Selenium WebDriver instance.
        by: Locator strategy (By.XPATH, By.ID, etc.).
        locator: The actual locator value.
        class_name: The class name to check for.
        timeout: Time to wait for the element (default 20 seconds).

    Returns:
        True if class name exists, False otherwise.
    """
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, locator))
        )
        classes = element.get_attribute("class")
        if classes:
            return class_name in classes.split()
        else:
            return False
    except TimeoutException:
        return False
    

def get_element_text(driver, by, locator, timeout=20):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, locator))
        )
        return element.text
    except TimeoutException:
        return None
    

def normalize_description_quotes(description):
    # Replace values inside double quotes with single quotes
    return re.sub(r'"([^"]+)"', r"'\1'", description)

def extract_value_and_operator_from_description(description):
    description = normalize_description_quotes(description)
    pattern = r"Enter a filter value of '(.+)' on the '(.+)' field using the '(.+)' filter operator."
    match = re.match(pattern, description)
    if match:
        value = match.group(1)
        field_name = match.group(2)
        operator = match.group(3)
        return {"value":value, "operator":operator, "field_name":field_name}
    else:
        return None, None
    

# def extract_multiple_values(description):
#     description = normalize_description_quotes(description)
#     pattern = r"Enter a filter value of '(.+)' on the '(.+)' field using the '(.+)' filter operator."
#     match = re.match(pattern, description)
#     if match:
#         value = match.group(1)
#         values_list = [v.strip() for v in value.split('/')]
#         return values_list
#     else:
#         return None
    
def extract_multiple_values(value_string):
    return [v.strip() for v in value_string.split('/')]

def extract_dates(date_str):
    parts = date_str.split('/')
    if len(parts) != 6:
        raise ValueError("Input string must contain two dates in the format dd/mm/yyyy/dd/mm/yyyy")
    
    from_date = f"{parts[0]}/{parts[1]}/{parts[2]}"
    to_date = f"{parts[3]}/{parts[4]}/{parts[5]}"
    
    return [from_date, to_date]


def click_back_button(driver, by, base_xpath,timeout=10):
    for i in range(1,100):
        # print("Attempt:", i)
        xpath = f"({base_xpath})[{i}]"
        try:
            close_button = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((by, xpath))
            )
            close_button.click()
            print(f"Successfully clicked button at index {i}")
            break
        except Exception as e:
            print(f"Attempt {i} failed for xpath : {xpath}")


def check_input_ancestor_is_table(driver, by, value_xpath, timeout=10):
    """
    Check if the input element (given by value_xpath) has a visible ancestor 
    with a specific class (e.g. part of a fixed data table row).
    """
    try:
        ancestor_xpath = f"{value_xpath}/ancestor::div[contains(@class,'fixedDataTableRowLayout_rowWrapper')]"
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, ancestor_xpath))
        )
        return element.is_displayed()
    except TimeoutException:
        return False
   

def extract_quickfilter_value(description):
    match = re.search(r"with a value of '([^']+)'", description)
    if match:
        return match.group(1)
    return None


def scroll_and_click_row(driver, by, container_xpath, target_xpath, timeout=10, max_scrolls=1000):
    time.sleep(2)
    # Try multiple container elements (if indexed)
    for i in range(1, 10):
        indexed_xpath = f"({container_xpath})[{i}]"
        try:
            container = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, indexed_xpath))
            )
            
            if container.is_displayed():
                print(f"Container found: {indexed_xpath}")
                actions = ActionChains(driver)
                # Determine scroll direction
                count = container.get_attribute("aria-rowcount")
                last_row_xpath = f"//div[contains(@class,'fixedDataTableRowLayout_')]/div[@aria-rowindex='{count}']"
                first_row_xpath = f"//div[contains(@class,'fixedDataTableRowLayout_')]/div[@aria-rowindex='2']"
                scroll_direction = None
                if check_element_exist(driver, by, last_row_xpath):
                    scroll_direction = Keys.PAGE_UP
                elif check_element_exist(driver, by, first_row_xpath):
                    scroll_direction = Keys.PAGE_DOWN

                for _ in range(max_scrolls):
                    try:
                        wait_and_click(driver, by, target_xpath)
                        print(f"Clicked element: {target_xpath}")
                        return
                    except TimeoutException:
                        actions.move_to_element(container).click().send_keys(scroll_direction).perform()
                        time.sleep(0.5)

                raise TimeoutException(f"Element {target_xpath} not found after scrolling.")

        except TimeoutException as e:
            print(f"Timeout or not displayed: {e}")
        except Exception as e:
            print(f"Unexpected error: {e}")

    print("Visible container not found.")

def scroll_into_view(driver, by, value, timeout=10):
    """
    Scrolls the specified element into view using JavaScript.
    Args:
        driver: Selenium WebDriver instance.
        by: Locator strategy (By.XPATH, By.ID, etc.).
        value: The actual locator value.
        timeout: Time to wait for the element (default 10 seconds).
    """
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, value))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", element)
        time.sleep(0.5)  # Optional wait after scrolling
         # Click the element after scrolling
    except TimeoutException:
        print(f"Element not found for scrolling: {value}")



def check_for_line_item_count(driver, by, item_number_xpath, timeout=10):
    try:
        item_number_count = WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((by, item_number_xpath))
        )
        count = len(item_number_count)
        print(f"Found {count} item(s).")
        return count
    except TimeoutException:
        print("Item elements not found within timeout.")
        return 0

def get_row_number_for_line_item(driver, by, line_item_container, total_items_count, timeout=10):
    line_number_xpath = line_item_container+"//input[contains(@aria-label,'Line number')]"
    item_number_xpath = line_item_container+"//input[contains(@aria-label,'Item number')]"
    try:
        for i in range(1, total_items_count + 1):
            line_indexed_xpath = f"({line_number_xpath})[{i}]"
            item_indexed_xpath = f"({item_number_xpath})[{i}]"
            print(line_indexed_xpath)
            print(item_indexed_xpath)
            try:
                item_element = WebDriverWait(driver, timeout).until(
                    EC.presence_of_element_located((by, line_indexed_xpath))
                )
                item_element_2 = WebDriverWait(driver, timeout).until(
                    EC.presence_of_element_located((by, item_indexed_xpath))
                )
                value_attr = item_element.get_attribute("value")
                value_attr_2 = item_element_2.get_attribute("value")
                if not value_attr and not value_attr_2:
                    return str(i)
                      # Return the row number of the first visible item
            except TimeoutException:
                continue
          # Not found
    except TimeoutException:
        print("Item elements not found within timeout.")


def check_element_has_child_elements(driver, by, element_xpath, timeout=10):
    try:
        parent_element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, element_xpath))
        )
        
        # Find child elements within the parent
        child_elements = parent_element.find_elements(By.XPATH, "./*")
        
        return len(child_elements)
    except TimeoutException:
        return False

def get_max_value_from_elements(driver, by, element_xpath, count, timeout=10):
    max_value = 0
    try:
        for i in range(1, count + 1):  # Include the last element
            indexed_xpath = f"({element_xpath})[{i}]"
            element = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by, indexed_xpath))
            )
            raw_value = element.get_attribute("value")
            try:
                value = int(raw_value)
                if value > max_value:
                    max_value = value
            except (ValueError, TypeError):
                print(f"Warning: Could not convert value '{raw_value}' to int at index {i}")
                continue

        return max_value
    except TimeoutException:
        print("Timeout while waiting for elements.")
        return None
    

def scroll_and_click_dropdown_item(driver, by ,container_xpath, target_locator, timeout=10, max_scrolls=30):
    try:
        container = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, container_xpath))
        )
 
        for _ in range(max_scrolls):
            try:
                target = WebDriverWait(driver, 1).until(
                    EC.element_to_be_clickable((by, target_locator))
                )
                target.click()
                print(f"✅ Clicked target: {target_locator}")
                return True
            except:
                # Scroll container using JavaScript
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].offsetHeight;", container)
                time.sleep(0.3)
 
        raise Exception(f"❌ Could not find target after scrolling: {target_locator}")
    except Exception as e:
        print(f"[scroll_and_click_dropdown_item] Error: {e}")
        return False


def check_if_checkbox_is_checked(driver, by, xpath, value, timeout=10):
    try:
        checkbox = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, xpath))
        )
        aria_checked = checkbox.get_attribute("aria-checked")
        if aria_checked == str(value).lower():
            return True
        else:
            return False
    except TimeoutException:
        return False
    
def check_aria_expanded(driver, by, xpath, timeout=10):
    try:
        checkbox = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, xpath))
        )
        aria_expanded = checkbox.get_attribute("aria-expanded")
        if aria_expanded == "false":
            return True
        else:
            return False
    except TimeoutException:
        return False
    
def remove_trailing_grid(text):
    if text.endswith("Grid"):
        return text[:-4]  # Remove last 4 characters
    return text

def scroll_and_click(driver, by, container_xpath, target_xpath, timeout=10, max_scrolls=100):
    time.sleep(2)
    # Try multiple container elements (if indexed)
    for i in range(1, 10):
        indexed_xpath = f"({container_xpath})[{i}]"
        try:
            container = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, indexed_xpath))
            )
           
            if container.is_displayed():
                print(f"Container found: {indexed_xpath}")
                actions = ActionChains(driver)
                scroll_direction = Keys.PAGE_DOWN
 
                for _ in range(max_scrolls):
                    try:
                        wait_and_click(driver, by, target_xpath)
                        print(f"Clicked element: {target_xpath}")
                        return
                    except TimeoutException:
                        actions.send_keys(scroll_direction).perform()
                        time.sleep(0.5)
 
                raise TimeoutException(f"Element {target_xpath} not found after scrolling.")
 
        except TimeoutException as e:
            print(f"Timeout or not displayed: {e}")
        except Exception as e:
            print(f"Unexpected error: {e}")
 
    print("Visible container not found.")
 

def get_element_attribute_value(driver, by, xpath, timeout=10):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, xpath))
        )
        return element.get_attribute("value")
    except TimeoutException:
        return None


# def take_screenshot(driver, name="screenshot"):
#     """
#     Takes a screenshot and saves it with a timestamp.
 
#     Args:
#         driver: Selenium WebDriver instance.
#         name: A name for the screenshot to be included in the filename.
#     """
#     try:
#         if not os.path.exists("screenshots"):
#             os.makedirs("screenshots")
#         timestamp = time.strftime("%Y%m%d_%H%M%S")
#         screenshot_name = f"{name}_{timestamp}.png"
#         screenshot_path = os.path.join("screenshots", screenshot_name)
#         driver.save_screenshot(screenshot_path)
#         print(f"Screenshot saved as {screenshot_path}")
#     except Exception as e:
#         print(f"Failed to take screenshot: {e}")
 
# def take_screenshot_on_pass(driver, test_name="test"):
#     """
#     Takes a screenshot on pass by calling the generic screenshot function.
 
#     Args:
#         driver: Selenium WebDriver instance.
#         test_name: A name for the test to be included in the filename.
#     """
#     take_screenshot(driver, f"passed_{test_name}")

# def take_screenshot_on_failure(driver, label="failure"):
#     timestamp = time.strftime("%Y%m%d-%H%M%S")
#     filename = f"screenshots/{label}_{timestamp}.png"
#     try:
#         driver.save_screenshot(filename)
#         print(f"📸 Screenshot taken: {filename}")
#     except Exception as e:
#         print(f"⚠️ Failed to take screenshot: {str(e)}")
 


import inspect

def get_script_name():
    """
    Inspects the call stack to return the actual test script file name that initiated execution.
    Skips over internal libraries (like pytest, runpy).
    """
    for frame in reversed(inspect.stack()):
        filename = frame.filename
        if filename.endswith(".py") and "site-packages" not in filename and "pytest" not in filename and "runpy" not in filename:
            return os.path.splitext(os.path.basename(filename))[0]
    return "unknown_script"

def take_screenshot(driver, name="screenshot", folder="screenshots"):
    """
    Takes a screenshot and saves it with a timestamp under the given folder.

    Args:
        driver: Selenium WebDriver instance.
        name: A name for the screenshot to be included in the filename.
        folder: Directory where the screenshot will be stored.
    """
    try:
        os.makedirs(folder, exist_ok=True)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        screenshot_name = f"{name}_{timestamp}.png"
        screenshot_path = os.path.join(folder, screenshot_name)
        driver.save_screenshot(screenshot_path)
        print(f"📸 Screenshot saved: {screenshot_path}")
    except Exception as e:
        print(f"⚠️ Failed to take screenshot: {e}")

def take_screenshot_on_pass(driver):
    """
    Takes a screenshot on test pass.
    """
    script_name = get_script_name()
    take_screenshot(driver, name=script_name, folder="reports/screens")

def take_screenshot_on_failure(driver):
    """
    Takes a screenshot on test failure.
    """
    script_name = get_script_name()
    take_screenshot(driver, name=script_name, folder="reports/screens")




def assert_navigation(driver,step_num="", *expected_items ):
    navigation_xpath = "//ol[@class='navigationBar-crumbList']//li"
    navigated_items = driver.find_elements(By.XPATH, navigation_xpath)
 
    navigated_items_texts = [item.text.strip() for item in navigated_items]
    print(f"Navigation found: {navigated_items_texts}")
 
    expected_path = list(expected_items)
    actual_path = navigated_items_texts[-len(expected_path):]
 
    if len(navigated_items_texts) < len(expected_path):
        take_screenshot_on_failure(driver)
        raise AssertionError(f"Not enough navigation items found.\nExpected at least {len(expected_path)}, got {len(navigated_items_texts)}\nNavigation: {navigated_items_texts}")
 
    if expected_path != actual_path:
        take_screenshot_on_failure(driver)
        raise AssertionError(f"mismatch.\nExpected: {expected_path}\nFound: {actual_path}")
    # desc = ""
    desc = " > ".join(expected_path)
    log_interaction(step_num, "Navigate", desc,"", "Pass")
    print("Navigation verified successfully.")
 
 
def click_nav(driver, by, base_xpath, timeout=10):
    global last_failed_xpath
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, base_xpath))
        )
        text = element.text.strip()
        aria_expanded = element.get_attribute("aria-expanded")
        if aria_expanded is None or aria_expanded.lower() == "false":
            WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, base_xpath)))
            ActionChains(driver).move_to_element(element).perform()
            element.click()
            # wait_and_click(driver, by, base_xpath, step_num, description, timeout, enable_fallback)
            # print(f"✅ Clicked element with base_xpath: {base_xpath}")
        return text if text else base_xpath
    except Exception as e:
        print(f"⚠️ Primary click failed on base_xpath: {base_xpath} - {str(e)}")
        last_failed_xpath = base_xpath
        take_screenshot_on_failure(driver)
        return False
 
# def validation(label,value,step_num="",description=""):
#     if value in description:
#         log_interaction(step_num, "Validation", label, value, "Pass")
#     elif value == "true" and "enabled" in description:
#         log_interaction(step_num, "Validation", label, value, "Pass")
#     elif value== "false" and "read" in description:
#         log_interaction(step_num, "Validation", label, value, "Pass")
#     else:
#         log_interaction(step_num, "Validation", label, value, "failed")


def func(a,b):
    if a < b:
        return a
    return b
    print(a)
    