
def read_all_test_data(filepath, sheetname="Sheet1"):
    import openpyxl, os
    abs_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', filepath))
    wb = openpyxl.load_workbook(abs_path)
    sheet = wb[sheetname]

    headers = [cell.value for cell in sheet[1]]
    all_data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):  # Skip empty rows
            continue
        row_dict = dict(zip(headers, row))
        all_data.append(row_dict)

    return all_data


# import openpyxl

# def read_all_test_data(file_path, sheet_name):
#     wb = openpyxl.load_workbook(file_path)
#     sheet = wb[sheet_name]
#     data = []
#     headers = [cell.value for cell in sheet[1]]

#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if any(row):
#             data.append(dict(zip(headers, row)))

#     return data




